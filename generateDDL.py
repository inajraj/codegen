from __future__ import print_function

import os.path
import mysql.connector as mariadb

import openpyxl

from dbUtils import runScript




def formFieldString(row):
    #get the row and form the field it could have the size or not a
    #also not null or not as well as primary key
    #all the int we need to automatically add unsigned
    #we also need to add auto_increment for the fields 'ID'
    finalStr = ""
    if (row[0]  == "ID"):
        finalStr = "`ID` "+ row[1] + ","
        return finalStr

    # need to see size is there or not
    

def RunDDLGenerator(wb, sheetId, CreateFlag, DataFlag):    

    #get the db credentials
    ws = wb['MetaData']
    
    #get the values from column H and I from row 1 to 4
    host = ws.cell(row=1,column=9).value
    user = ws.cell(2,9).value
    passwd = ws.cell(3,9).value
    database = ws.cell(4,9).value   
   
    print(host, user, passwd, database)
        
    mydb = mariadb.connect(
             host=host,
             user=user,
             passwd=passwd,
             database=database
         )

    #print (sheetId)
    ws = wb[sheetId]

    if (CreateFlag == 'Y'):
         #create the table only if the indicator is Y
         #drop the table first
        pkString = ""
        str = "DROP TABLE IF EXISTS " + sheetId
        runScript(str,mydb)
        finalStr = "CREATE TABLE `" + sheetId + "` ("
        i = 2
        while i <= ws.max_row:
            print(i)
            print(ws.cell(row=i,column=2).value)
            finalStr  = finalStr + "`" + ws.cell(row=i,column=2).value + "` " + ws.cell(row=i,column=3).value + ","
            #print(len(ws.cell(i,2).value))
            #check this row has primary key column
            if (ws.cell(row=i,column=4).value == 'PK'):
                pkString = "PRIMARY KEY (`" + ws.cell(row=i,column=2).value + "`))"
            i = i + 1
            #need to add primary key - to check the loop again
        runScript(finalStr+pkString, mydb)
            
        print(finalStr+pkString)
    if (DataFlag == 'Y'):
        #col 1 has tablename
        tableName = ws.cell(row=2,column=1).value
        basesql = "INSERT " + tableName
        #col 2 has fieldnames
        fields = "("
        i = 2
        while i <= ws.max_row:
            #skip ID
            if ws.cell(row=i,column=2).value != 'ID': #skip ID auto generated
                fields = fields + ws.cell(row=i,column=2).value + ","
            i = i+1
        fields = fields[:-1] + ")"
        print(fields)
        basesql = basesql + fields
        #need to generate insert statements for every DATA column
        #start looking for column header 'Data'
        col = 4 
        while col <= ws.max_column:
            print(col)
            if (ws.cell(row=1,column=col).value == 'Data'):
                values = "('"
                i = 2 
                while i <= ws.max_row:
                    if ws.cell(row=i,column=col).value != 'ID': #skip ID as it is auto generated
                        if ws.cell(row=i,column=col).value != None:
                            values = values + ws.cell(row=i,column=col).value + "','" 
                        else
                            values = values + 'NULL','"
                    i = i + 1
                values = values[:-2] + ")"  
                print (values)   
                sql = basesql + "  VALUES " + values
                print(sql)
                runScript(sql, mydb)
            col = col + 1    

    
def generateInsertStringforPhp(sheet,sheetId,rangeName):
    
    result = sheet.values().get(spreadsheetId=sheetId,
                                range=rangeName).execute()
    values = result.get('values', [])
    
    for metaRow in values:
        res=sheet.values().get(spreadsheetId=sheetId,
                                range=metaRow[0]).execute()
        tablerows = res.get('values', [])
      
        if (metaRow[3] == 'Y'):
            qStr = "\"insert into " + metaRow[0] + "("
            for tdRow in tablerows:
                if ("NOT NULL" in tdRow[1]):
                    qStr = qStr + tdRow[0] + ","
                else:
                    qStr = qStr +  "\" . checkBlank(\""+ tdRow[0] + "\", $results) . \""
            qStr = qStr.rstrip(',') + ")" 
            qStr = qStr + " Values ('\""
            for tdRow in tablerows:
                if ("NOT NULL" in tdRow[1]):
                    qStr = qStr + " . $results[\'" + tdRow[0] + "\'] . \"','\""
                else:
                    qStr = qStr[:-3] +  "\" . removeBlank($results[\'" + tdRow[0] + "\']) . \" ,'\""

            qStr = qStr[:-5] + "\")\"" 
            print(qStr)

def generateUpdateStringforPhp(sheet,sheetId,rangeName):
    
    result = sheet.values().get(spreadsheetId=sheetId,
                                range=rangeName).execute()
    values = result.get('values', [])
    
    for metaRow in values:
        res=sheet.values().get(spreadsheetId=sheetId,
                                range=metaRow[0]).execute()
        tablerows = res.get('values', [])
        whereStr = ""
        if (metaRow[3] == 'Y'):
            qStr = "\"update " + metaRow[0] + " SET "
            for tdRow in tablerows:
                if ("PK" in tdRow[2]):
                    whereStr = ". \" where " + tdRow[0] + " = '\" . $results[\'" + tdRow[0] + "\'] . \"'\" "
                if ("NOT NULL" in tdRow[1]):
                    qStr = qStr  + tdRow[0] +  " = '\" . $results[\'" + tdRow[0] + "\'] . \"', "
                else:
                    qStr = qStr +  "\" . checkBlank(" + tdRow[0] + ", $results[\'" + tdRow[0] + "\']) .\" "
            qStr = qStr[:-3] + "'\" " + whereStr
            
            print(qStr)

     