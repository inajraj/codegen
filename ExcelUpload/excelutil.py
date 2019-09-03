import  openpyxl 
from openpyxl import load_workbook
import json

from  References import LCDict, ListColumnsDict


def insertMyCommunity(tVal, Colval):
    
    #get the number in the end of community
    num = tVal[-1:]
    print( "dafadsf" + num)


    if Colval is None:
        LCdata['CommunityName'+num] = ''
        LCdata['CommunityMobile'+num] = ''
        LCdata['CommunityEmail'+num] = ''
    else:
        MyCom = Colval.split(',')
        LCdata['CommunityName'+num] = MyCom[0]
        LCdata['CommunityMobile'+num] = MyCom[1]
        LCdata['CommunityEmail'+num] = MyCom[2]
    #split the column val
    




wb = load_workbook('LCData.xlsx')

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Data')
NoOfColumns = 0
for i in range(1,50):
    if (sheet.cell(row=1,column=i).value is None):
        NoOfColumns = i - 1
        break
    print(i, sheet.cell(row=1,column=i).value)
print(NoOfColumns)

NoOfRows = 0
for i in range(2,1000):
    if (sheet.cell(row=i,column=1).value is None):
        NoOfRows = i - 2
        break
print(NoOfRows)

LCdata = {}

for rowId in range (3, 4):
    for ColId in range(2, NoOfColumns):
        
        tVal = sheet.cell(row=1,column=ColId).value
        print( tVal)
       

        if tVal in LCDict.keys():
            LCdata[LCDict[tVal]] = sheet.cell(row=rowId,column=ColId).value
        else:
            # key is not there but the value will match -- so get the key name from the value
            if tVal in ListColumnsDict.keys():
                LCdata[ListColumnsDict[tVal]] = sheet.cell(row=rowId,column=ColId).value
         #for stream of education alone we need to pick up others column
        if 'StreamofEducation' in tVal:
            
            if ('Others' in sheet.cell(row=rowId,column=ColId).value):
                
                LCdata['StreamofEducation'] =  sheet.cell(row=rowId,column=ColId+1).value
            else:
                LCdata['StreamofEducation'] =  sheet.cell(row=rowId,column=ColId).value

        if 'Community' in tVal:
            insertMyCommunity(tVal, sheet.cell(row=rowId,column=ColId).value)

    json_data = json.dumps(LCdata)
print(LCdata)